#!/usr/bin/env python3

import textract
import os, re
from openpyxl import Workbook

path = "/path"

list_requirementes = [

]

wb = Workbook()
ws = wb.active
ws.title = "Extracted Info"


# REMOVE SPACES
def remove_space(name):
    return name.replace(" ", "")

# REMOVE SPACES
def fix_name_doc():
    files_name = []
    for filename in os.listdir(path):
        if " " in filename:
            fixed_name = remove_space(filename)

            src = os.path.join(path, filename)
            dst = os.path.join(path, fixed_name)

            # Only rename if names are different
            if filename != fixed_name:
                os.rename(src, dst)
                print("FIXED: " + filename + " → " + fixed_name)

        files_name.append(filename)

    return files_name

# AUXILIAR REGEX FOR BEFORE
def extract_before_header(text):
    pattern = re.compile(r'((?:[A-ZÁÉÍÓÚÑÜ]{2,}\s+){4,}[A-ZÁÉÍÓÚÑÜ]{2,})')
    match = pattern.search(text)
    if match:
        return text[:match.start()].strip()
    else:
        return text.strip()


# MAIN REGEX TO FIND THE MATCH
def get_info(keyword):
    files_name = fix_name_doc()

    for file in files_name:
        full_path = path + file
        print(f"Processing file: {file}")

        try:
            text_bytes = textract.process(full_path)
            text = text_bytes.decode('utf-8', errors='ignore')
        except Exception as e:
            print(f"Error reading {file}: {e}")
            continue

        normalized_text = re.sub(r'\s+', ' ', text).strip()

        pattern = re.compile(rf"{re.escape(keyword)}\s*[:\-]?\s*(.+)")
        match = pattern.search(normalized_text)

        if match:
            project_big_text = match.group(1).strip()
            clean_project_name = extract_before_header(project_big_text)
            print(f"Found in {file}: {clean_project_name}")
            return clean_project_name

        print(f"No match found in {file}")

    return ""

if not os.path.exists("/path/extracted_info.xlsx"):
    # Write column headers (requirement names) in first row (A1, B1, ...)
    for col_index, header in enumerate(list_requirementes, start=1):
        ws.cell(row=1, column=col_index, value=header)

def main():

    list_result = []

    for requirement in list_requirementes:
        result = get_info(requirement)
        list_result.append(result)


    # Write headers in first row
    for col_index, header in enumerate(list_requirementes, start=1):
        ws.cell(row=1, column=col_index, value=header)

    # Write values in second row under each header
    for col_index, value in enumerate(list_result, start=1):
        ws.cell(row=2, column=col_index, value=value)

    print(list_result)
    wb.save("extracted_info.xlsx")
    

if __name__ == "__main__":
    main()

